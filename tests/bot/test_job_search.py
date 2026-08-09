"""src/bot/job_search.py 單元測試（對應 robinson SPEC.md FR-33～FR-38，Step 4.1／4.2，ADR-24／ADR-26）。"""
import io
from datetime import date, datetime
from unittest.mock import MagicMock
from zoneinfo import ZoneInfo

import openpyxl

from src.bot import job_search

_TAIWAN_TZ = ZoneInfo("Asia/Taipei")

# --- is_text_length_valid（FR-36）---


def test_is_text_length_valid_within_limit():
    assert job_search.is_text_length_valid("我的履歷內容")


def test_is_text_length_valid_at_exact_limit():
    assert job_search.is_text_length_valid("A" * 3500)


def test_is_text_length_valid_over_limit():
    assert not job_search.is_text_length_valid("A" * 3501)


# --- is_years_of_experience_reasonable（FR-36，ADR-26 決策 1）---


def test_is_years_of_experience_reasonable_zero_is_valid():
    assert job_search.is_years_of_experience_reasonable(0)


def test_is_years_of_experience_reasonable_typical_value():
    assert job_search.is_years_of_experience_reasonable(3.5)


def test_is_years_of_experience_reasonable_boundary_max():
    assert job_search.is_years_of_experience_reasonable(60)


def test_is_years_of_experience_reasonable_over_max():
    assert not job_search.is_years_of_experience_reasonable(60.1)


def test_is_years_of_experience_reasonable_negative():
    assert not job_search.is_years_of_experience_reasonable(-1)


# --- save_search_criteria（FR-33，ADR-24 決策 3：允許同時存多組）---


def test_save_search_criteria_inserts_row(fake_db):
    criteria_id = job_search.save_search_criteria(fake_db, 1, "AI 工程師", None, 50000, None)

    row = fake_db.select("job_search_criteria", where="id = %s", params=(criteria_id,), fetch_one=True)
    assert row["user_id"] == 1
    assert row["keyword"] == "AI 工程師"
    assert row["region"] is None
    assert row["salary_min"] == 50000
    assert row["salary_max"] is None


def test_save_search_criteria_allows_multiple_rows_for_same_user(fake_db):
    job_search.save_search_criteria(fake_db, 1, "AI 工程師", None, 50000, None)
    job_search.save_search_criteria(fake_db, 1, "資料工程師", "台北", None, None)

    rows = fake_db.select("job_search_criteria", where="user_id = %s", params=(1,))
    assert len(rows) == 2
    assert {row["keyword"] for row in rows} == {"AI 工程師", "資料工程師"}


# --- save_profile（FR-36）---


def test_save_profile_updates_user_fields(fake_db):
    user_id = fake_db.insert("users", {"telegram_user_id": 1, "role": "Robin", "is_owner": True})

    job_search.save_profile(fake_db, user_id, "履歷內容", "期望工作內容", 3.5, 50000, 70000)

    row = fake_db.select("users", where="id = %s", params=(user_id,), fetch_one=True)
    assert row["job_resume"] == "履歷內容"
    assert row["job_expectation"] == "期望工作內容"
    assert row["years_of_experience"] == 3.5
    assert row["expected_salary_min"] == 50000
    assert row["expected_salary_max"] == 70000


# --- list_search_criteria（FR-33）---


def test_list_search_criteria_returns_only_this_user(fake_db):
    job_search.save_search_criteria(fake_db, 1, "AI 工程師", None, None, None)
    job_search.save_search_criteria(fake_db, 2, "別人的條件", None, None, None)

    rows = job_search.list_search_criteria(fake_db, 1)

    assert len(rows) == 1
    assert rows[0]["keyword"] == "AI 工程師"


# --- upsert_company（FR-35a）---


def test_upsert_company_inserts_new_company_with_null_background(fake_db):
    company_id, is_new = job_search.upsert_company(fake_db, "999", "某某科技", "台北市信義區", "軟體業")

    assert is_new is True
    row = fake_db.select("job_companies", where="id = %s", params=(company_id,), fetch_one=True)
    assert row["company_id_104"] == "999"
    assert row["company_name"] == "某某科技"
    assert row["background"] is None


def test_upsert_company_existing_company_does_not_overwrite(fake_db):
    existing_id = fake_db.insert(
        "job_companies",
        {"company_id_104": "999", "company_name": "某某科技", "region": "台北市", "background": "已回填的背景"},
    )

    company_id, is_new = job_search.upsert_company(fake_db, "999", "改名後的公司", "高雄市", "製造業")

    assert is_new is False
    assert company_id == existing_id
    row = fake_db.select("job_companies", where="id = %s", params=(existing_id,), fetch_one=True)
    assert row["company_name"] == "某某科技"
    assert row["background"] == "已回填的背景"


# --- upsert_job_posting（FR-34d）---


_JOB = {
    "job_id": "12345",
    "job_slug": "abcde",
    "title": "AI 工程師",
    "company_id": "999",
    "company_name": "某某科技",
    "region": "台北市信義區",
    "url": "https://www.104.com.tw/job/12345",
    "applicant_count": None,
    "is_closed": False,
}
_DETAIL = {
    "content": "負責 AI 模型開發",
    "required_years_experience": 3.0,
    "source_updated_at": None,
}


def test_upsert_job_posting_inserts_new_job(fake_db):
    now = datetime(2026, 8, 10, 8, 0, tzinfo=_TAIWAN_TZ)

    is_new = job_search.upsert_job_posting(fake_db, _JOB, _DETAIL, now)

    assert is_new is True
    row = fake_db.select("job_postings", where="job_id_104 = %s", params=("12345",), fetch_one=True)
    assert row["title"] == "AI 工程師"
    assert row["content"] == "負責 AI 模型開發"
    assert row["required_years_experience"] == 3.0
    assert row["first_seen_at"] == now
    assert row["last_crawled_at"] == now
    assert row["is_closed"] is False


def test_upsert_job_posting_updates_existing_job_without_changing_first_seen_at(fake_db):
    first_seen = datetime(2026, 8, 3, 8, 0, tzinfo=_TAIWAN_TZ)
    fake_db.insert(
        "job_postings",
        {
            "job_id_104": "12345", "company_id_104": "999", "title": "舊職缺名稱", "region": "台北市",
            "url": "https://www.104.com.tw/job/12345", "content": "舊內容",
            "required_years_experience": None, "applicant_count": None, "source_updated_at": None,
            "first_seen_at": first_seen, "last_crawled_at": first_seen, "is_closed": False,
        },
    )
    second_crawl = datetime(2026, 8, 10, 8, 0, tzinfo=_TAIWAN_TZ)

    is_new = job_search.upsert_job_posting(fake_db, _JOB, _DETAIL, second_crawl)

    assert is_new is False
    row = fake_db.select("job_postings", where="job_id_104 = %s", params=("12345",), fetch_one=True)
    assert row["title"] == "AI 工程師"
    assert row["content"] == "負責 AI 模型開發"
    assert row["first_seen_at"] == first_seen
    assert row["last_crawled_at"] == second_crawl


def test_upsert_job_posting_updates_is_closed_status_on_recrawl(fake_db):
    """職缺重新爬到時，`is_closed` 要跟著更新（例如原本開放的職缺後來關閉了），不能卡在第一次
    爬到時的舊狀態（見 `upsert_job_posting()` docstring）。"""
    first_seen = datetime(2026, 8, 3, 8, 0, tzinfo=_TAIWAN_TZ)
    fake_db.insert(
        "job_postings",
        {
            "job_id_104": "12345", "company_id_104": "999", "title": "AI 工程師", "region": "台北市",
            "url": "https://www.104.com.tw/job/12345", "content": "舊內容",
            "required_years_experience": None, "applicant_count": None, "source_updated_at": None,
            "first_seen_at": first_seen, "last_crawled_at": first_seen, "is_closed": False,
        },
    )
    closed_job = {**_JOB, "is_closed": True}
    second_crawl = datetime(2026, 8, 10, 8, 0, tzinfo=_TAIWAN_TZ)

    job_search.upsert_job_posting(fake_db, closed_job, _DETAIL, second_crawl)

    row = fake_db.select("job_postings", where="job_id_104 = %s", params=("12345",), fetch_one=True)
    assert row["is_closed"] is True


# --- crawl_and_upsert_jobs（FR-34a～FR-34d）---


class _FakeJob104Client:
    """模擬 submodules.job104.client.Job104Client，依 (keyword, page) 回傳預先設定好的清單，
    依 job_slug（詳情頁 API 用的短代碼 ID，跟 job_id/jobNo 是不同識別碼，2026-08-09 實測驗證
    確認，見 submodules/job104/client.py 模組 docstring）回傳預先設定好的詳情，並記錄呼叫
    次數供測試驗證分頁/延遲行為。"""

    def __init__(self, pages: dict, details: dict):
        self._pages = pages
        self._details = details
        self.search_calls: list[tuple] = []
        self.detail_calls: list[str] = []

    def search_list(self, keyword, salary_min=None, salary_max=None, page=1):
        self.search_calls.append((keyword, page))
        return self._pages.get((keyword, page), [])

    def fetch_job_detail(self, job_slug):
        self.detail_calls.append(job_slug)
        return self._details.get(job_slug, {})


def _fake_sleep(_seconds):
    pass


def _fake_random(_a, _b):
    return 0


def test_crawl_and_upsert_jobs_no_criteria_makes_no_requests(fake_db):
    client = _FakeJob104Client(pages={}, details={})

    result = job_search.crawl_and_upsert_jobs(fake_db, client, 1, sleep_func=_fake_sleep, random_func=_fake_random)

    assert result == {"new_company_ids": [], "new_job_count": 0, "updated_job_count": 0}
    assert client.search_calls == []


def test_crawl_and_upsert_jobs_single_criteria_single_page(fake_db):
    job_search.save_search_criteria(fake_db, 1, "AI 工程師", "台北市", 50000, None)
    page1 = [
        {"job_id": "1", "job_slug": "slug1", "title": "AI 工程師 A", "company_id": "100",
         "company_name": "A 公司", "region": "台北市", "url": "https://www.104.com.tw/job/1",
         "applicant_count": 3},
        {"job_id": "2", "job_slug": "slug2", "title": "AI 工程師 B", "company_id": "100",
         "company_name": "A 公司", "region": "台北市", "url": "https://www.104.com.tw/job/2",
         "applicant_count": 5},
    ]
    client = _FakeJob104Client(
        pages={("AI 工程師", 1): page1},
        details={"slug1": {"content": "內容1"}, "slug2": {"content": "內容2"}},
    )

    result = job_search.crawl_and_upsert_jobs(fake_db, client, 1, sleep_func=_fake_sleep, random_func=_fake_random)

    assert result == {"new_company_ids": ["100"], "new_job_count": 2, "updated_job_count": 0}
    assert client.search_calls == [("AI 工程師", 1), ("AI 工程師", 2)]  # 第 2 頁空清單才停止翻頁
    assert sorted(client.detail_calls) == ["slug1", "slug2"]
    jobs = fake_db.select("job_postings")
    assert len(jobs) == 2
    companies = fake_db.select("job_companies")
    assert len(companies) == 1


def test_crawl_and_upsert_jobs_filters_by_region_substring(fake_db):
    """2026-08-09 依 Robin 指示：地區篩選不送給 104 API（`area` 需要數字代碼，沒有可靠對照
    表），改由呼叫端對回傳結果的 `region` 文字做子字串比對篩選（見 `crawl_and_upsert_jobs()`
    docstring）。"""
    job_search.save_search_criteria(fake_db, 1, "AI 工程師", "台北市", None, None)
    page1 = [
        {"job_id": "1", "job_slug": "slug1", "title": "符合地區", "company_id": "100",
         "company_name": "A 公司", "region": "台北市信義區", "url": "https://www.104.com.tw/job/1",
         "applicant_count": 1},
        {"job_id": "2", "job_slug": "slug2", "title": "不符合地區", "company_id": "200",
         "company_name": "B 公司", "region": "新竹市", "url": "https://www.104.com.tw/job/2",
         "applicant_count": 2},
    ]
    client = _FakeJob104Client(
        pages={("AI 工程師", 1): page1}, details={"slug1": {}, "slug2": {}},
    )

    result = job_search.crawl_and_upsert_jobs(fake_db, client, 1, sleep_func=_fake_sleep, random_func=_fake_random)

    assert result["new_job_count"] == 1
    assert client.detail_calls == ["slug1"]  # 不符合地區的 slug2 完全不會呼叫詳情頁
    jobs = fake_db.select("job_postings")
    assert [j["title"] for j in jobs] == ["符合地區"]


def test_crawl_and_upsert_jobs_region_filter_does_not_stop_pagination_early(fake_db):
    """單頁篩選後一筆都不剩，也不能誤判成「這組條件已經爬完」而提早停止翻頁——分頁停止判斷
    要看 104 回傳的原始清單是否為空，不是篩選後的清單。"""
    job_search.save_search_criteria(fake_db, 1, "AI 工程師", "台北市", None, None)
    page1 = [
        {"job_id": "1", "job_slug": "slug1", "title": "不符合地區", "company_id": "100",
         "company_name": "A 公司", "region": "新竹市", "url": "https://www.104.com.tw/job/1",
         "applicant_count": 1},
    ]
    client = _FakeJob104Client(pages={("AI 工程師", 1): page1, ("AI 工程師", 2): []}, details={})

    result = job_search.crawl_and_upsert_jobs(fake_db, client, 1, sleep_func=_fake_sleep, random_func=_fake_random)

    assert result == {"new_company_ids": [], "new_job_count": 0, "updated_job_count": 0}
    assert client.search_calls == [("AI 工程師", 1), ("AI 工程師", 2)]
    assert client.detail_calls == []


def test_crawl_and_upsert_jobs_stops_pagination_on_empty_page(fake_db):
    job_search.save_search_criteria(fake_db, 1, "AI 工程師", None, None, None)
    client = _FakeJob104Client(pages={("AI 工程師", 1): []}, details={})

    result = job_search.crawl_and_upsert_jobs(fake_db, client, 1, sleep_func=_fake_sleep, random_func=_fake_random)

    assert result == {"new_company_ids": [], "new_job_count": 0, "updated_job_count": 0}
    assert client.search_calls == [("AI 工程師", 1)]


def test_crawl_and_upsert_jobs_multiple_criteria_each_queried(fake_db):
    job_search.save_search_criteria(fake_db, 1, "AI 工程師", None, None, None)
    job_search.save_search_criteria(fake_db, 1, "資料工程師", None, None, None)
    client = _FakeJob104Client(
        pages={
            ("AI 工程師", 1): [{"job_id": "1", "job_slug": "slug1", "title": "A", "company_id": "100",
                              "company_name": "A 公司", "region": "台北市",
                              "url": "https://www.104.com.tw/job/1", "applicant_count": 1}],
            ("資料工程師", 1): [{"job_id": "2", "job_slug": "slug2", "title": "B", "company_id": "200",
                             "company_name": "B 公司", "region": "新竹市",
                             "url": "https://www.104.com.tw/job/2", "applicant_count": 2}],
        },
        details={"slug1": {}, "slug2": {}},
    )

    result = job_search.crawl_and_upsert_jobs(fake_db, client, 1, sleep_func=_fake_sleep, random_func=_fake_random)

    assert result["new_job_count"] == 2
    assert sorted(result["new_company_ids"]) == ["100", "200"]


def test_crawl_and_upsert_jobs_applies_delay_between_every_request(fake_db):
    job_search.save_search_criteria(fake_db, 1, "AI 工程師", None, None, None)
    client = _FakeJob104Client(
        pages={("AI 工程師", 1): [{"job_id": "1", "job_slug": "slug1", "title": "A", "company_id": "100",
                                 "company_name": "A 公司", "region": "台北市",
                                 "url": "https://www.104.com.tw/job/1", "applicant_count": 1}]},
        details={"slug1": {}},
    )
    delay_calls = []

    def _tracking_sleep(seconds):
        delay_calls.append(seconds)

    job_search.crawl_and_upsert_jobs(
        fake_db, client, 1, sleep_func=_tracking_sleep, random_func=lambda a, b: 2.5
    )

    # 3 次請求：page1 列表、job1 詳情、page2 列表（空清單才停止），每次請求後都要延遲一次。
    assert len(client.search_calls) == 2
    assert len(client.detail_calls) == 1
    assert delay_calls == [2.5, 2.5, 2.5]
    for seconds in delay_calls:
        assert 2 <= seconds <= 4


# --- FR-35：公司背景 Email／CSV／Drive 人力協作機制 ---


def test_company_csv_filename():
    assert job_search.company_csv_filename(date(2026, 8, 9)) == "2026-08-09-104職缺公司.csv"


def test_get_companies_by_ids_returns_matching_rows_and_skips_missing(fake_db):
    fake_db.insert(
        "job_companies", {"company_id_104": "100", "company_name": "A 公司", "region": "台北市", "background": None}
    )
    fake_db.insert(
        "job_companies", {"company_id_104": "200", "company_name": "B 公司", "region": "新竹市", "background": None}
    )

    rows = job_search.get_companies_by_ids(fake_db, ["100", "999", "200"])

    assert [r["company_id_104"] for r in rows] == ["100", "200"]


def test_build_new_companies_csv_has_header_and_empty_background_column():
    companies = [
        {"company_id_104": "100", "company_name": "A 公司", "region": "台北市", "industry": "軟體業"},
        {"company_id_104": "200", "company_name": "B 公司", "region": None, "industry": None},
    ]

    csv_text = job_search.build_new_companies_csv(companies)

    lines = csv_text.strip().splitlines()
    assert lines[0] == "104公司ID,公司全名,地區,產業類型,背景"
    assert lines[1] == "100,A 公司,台北市,軟體業,"
    assert lines[2] == "200,B 公司,,,"


def test_send_new_companies_email_calls_client_with_expected_args():
    email_client = MagicMock()

    job_search.send_new_companies_email(email_client, "robin@gmail.com", date(2026, 8, 9), "104公司ID,背景\n100,\n")

    email_client.send_text_with_attachment.assert_called_once()
    call_kwargs = email_client.send_text_with_attachment.call_args.kwargs
    assert call_kwargs["to"] == "robin@gmail.com"
    assert call_kwargs["subject"] == "2026-08-09 排程 - Robinson 104 職缺公司列表"
    assert call_kwargs["body"] == "附件為本週爬到的最新公司列表，請參閱！"
    assert call_kwargs["attachment_filename"] == "2026-08-09-104職缺公司.csv"
    assert call_kwargs["attachment_bytes"] == "104公司ID,背景\n100,\n".encode("utf-8-sig")


def test_parse_companies_csv_extracts_filled_rows():
    csv_text = "104公司ID,公司全名,地區,產業類型,背景\n100,A 公司,台北市,軟體業,做電商平台的新創\n"

    entries = job_search.parse_companies_csv(csv_text)

    assert entries == [{"company_id_104": "100", "background": "做電商平台的新創"}]


def test_parse_companies_csv_skips_rows_with_empty_background():
    csv_text = "104公司ID,公司全名,地區,產業類型,背景\n100,A 公司,台北市,軟體業,\n"

    assert job_search.parse_companies_csv(csv_text) == []


def test_parse_companies_csv_skips_rows_with_empty_company_id():
    csv_text = "104公司ID,公司全名,地區,產業類型,背景\n,A 公司,台北市,軟體業,某個背景\n"

    assert job_search.parse_companies_csv(csv_text) == []


def test_apply_company_backgrounds_updates_matching_and_reports_not_found(fake_db):
    fake_db.insert(
        "job_companies", {"company_id_104": "100", "company_name": "A 公司", "region": "台北市", "background": None}
    )
    entries = [
        {"company_id_104": "100", "background": "電商新創"},
        {"company_id_104": "999", "background": "不存在的公司"},
    ]

    result = job_search.apply_company_backgrounds(fake_db, entries)

    assert result == {"updated_count": 1, "not_found_ids": ["999"]}
    row = fake_db.select("job_companies", where="company_id_104 = %s", params=("100",), fetch_one=True)
    assert row["background"] == "電商新創"


# --- check_and_run_weekly_job_search（FR-34b、FR-35a～FR-35c）---

_MONDAY_8AM = datetime(2026, 8, 10, 8, 0, tzinfo=_TAIWAN_TZ)  # 2026-08-10 是週一


def _seed_owner(fake_db, job_search_last_run_on=None):
    return fake_db.insert(
        "users",
        {
            "telegram_user_id": 8263904025, "role": "Robin", "is_owner": True,
            "job_search_last_run_on": job_search_last_run_on,
        },
    )


def test_check_and_run_weekly_job_search_skips_when_not_monday_8am(fake_db):
    _seed_owner(fake_db)
    job_search.save_search_criteria(fake_db, 1, "AI 工程師", None, None, None)
    email_client = MagicMock()
    telegram_client = MagicMock()
    not_monday = datetime(2026, 8, 11, 8, 0, tzinfo=_TAIWAN_TZ)

    job_search.check_and_run_weekly_job_search(
        fake_db, MagicMock(), email_client, "robin@gmail.com", telegram_client, now=not_monday
    )

    email_client.send_text_with_attachment.assert_not_called()
    telegram_client.send_text.assert_not_called()


def test_check_and_run_weekly_job_search_skips_when_no_owner(fake_db):
    email_client = MagicMock()
    telegram_client = MagicMock()

    job_search.check_and_run_weekly_job_search(
        fake_db, MagicMock(), email_client, "robin@gmail.com", telegram_client, now=_MONDAY_8AM
    )

    email_client.send_text_with_attachment.assert_not_called()


def test_check_and_run_weekly_job_search_skips_when_toggle_disabled(fake_db):
    owner_id = _seed_owner(fake_db)
    fake_db.insert("feature_toggles", {"user_id": owner_id, "feature_key": "job_search", "is_enabled": False})
    email_client = MagicMock()
    telegram_client = MagicMock()

    job_search.check_and_run_weekly_job_search(
        fake_db, MagicMock(), email_client, "robin@gmail.com", telegram_client, now=_MONDAY_8AM
    )

    email_client.send_text_with_attachment.assert_not_called()


def test_check_and_run_weekly_job_search_skips_when_already_run_today(fake_db):
    _seed_owner(fake_db, job_search_last_run_on=_MONDAY_8AM.date())
    email_client = MagicMock()
    telegram_client = MagicMock()

    job_search.check_and_run_weekly_job_search(
        fake_db, MagicMock(), email_client, "robin@gmail.com", telegram_client, now=_MONDAY_8AM
    )

    email_client.send_text_with_attachment.assert_not_called()


def test_check_and_run_weekly_job_search_no_new_companies_skips_email(fake_db, monkeypatch):
    monkeypatch.setattr(job_search, "_polite_delay", lambda *a, **k: None)
    owner_id = _seed_owner(fake_db)
    fake_db.insert(
        "job_companies", {"company_id_104": "100", "company_name": "A 公司", "region": "台北市", "background": "已知背景"}
    )
    job_search.save_search_criteria(fake_db, owner_id, "AI 工程師", None, None, None)
    client = _FakeJob104Client(
        pages={
            ("AI 工程師", 1): [
                {"job_id": "1", "job_slug": "slug1", "title": "A", "company_id": "100",
                 "company_name": "A 公司", "region": "台北市", "url": "https://www.104.com.tw/job/1",
                 "applicant_count": 1}
            ]
        },
        details={"slug1": {}},
    )
    email_client = MagicMock()
    telegram_client = MagicMock()

    job_search.check_and_run_weekly_job_search(
        fake_db, client, email_client, "robin@gmail.com", telegram_client, now=_MONDAY_8AM
    )

    email_client.send_text_with_attachment.assert_not_called()
    telegram_client.send_text.assert_not_called()
    owner_row = fake_db.select("users", where="id = %s", params=(owner_id,), fetch_one=True)
    assert owner_row["job_search_last_run_on"] == _MONDAY_8AM.date()


def test_check_and_run_weekly_job_search_new_companies_sends_email_and_notifies(fake_db, monkeypatch):
    monkeypatch.setattr(job_search, "_polite_delay", lambda *a, **k: None)
    owner_id = _seed_owner(fake_db)
    job_search.save_search_criteria(fake_db, owner_id, "AI 工程師", None, None, None)
    client = _FakeJob104Client(
        pages={
            ("AI 工程師", 1): [
                {"job_id": "1", "job_slug": "slug1", "title": "A", "company_id": "100",
                 "company_name": "A 公司", "region": "台北市", "url": "https://www.104.com.tw/job/1",
                 "applicant_count": 1}
            ]
        },
        details={"slug1": {}},
    )
    email_client = MagicMock()
    telegram_client = MagicMock()

    job_search.check_and_run_weekly_job_search(
        fake_db, client, email_client, "robin@gmail.com", telegram_client, now=_MONDAY_8AM
    )

    email_client.send_text_with_attachment.assert_called_once()
    call_kwargs = email_client.send_text_with_attachment.call_args.kwargs
    assert call_kwargs["to"] == "robin@gmail.com"
    assert "100,A 公司" in call_kwargs["attachment_bytes"].decode("utf-8-sig")
    telegram_client.send_text.assert_called_once_with(
        chat_id=8263904025, text=job_search.EMAIL_SENT_NOTIFICATION_TEXT
    )
    owner_row = fake_db.select("users", where="id = %s", params=(owner_id,), fetch_one=True)
    assert owner_row["job_search_last_run_on"] == _MONDAY_8AM.date()


# --- list_scorable_jobs（FR-37a）---


def _insert_company(fake_db, company_id_104, background="已回填的背景", **overrides):
    row = {
        "company_id_104": company_id_104, "company_name": f"{company_id_104} 公司", "region": "台北市",
        "industry": "軟體業", "background": background,
    }
    row.update(overrides)
    return fake_db.insert("job_companies", row)


def _insert_job(fake_db, job_id_104, company_id_104, **overrides):
    row = {
        "job_id_104": job_id_104, "company_id_104": company_id_104, "title": f"職缺 {job_id_104}",
        "region": "台北市", "url": f"https://www.104.com.tw/job/{job_id_104}", "content": "職缺內容",
        "required_years_experience": 3.0, "applicant_count": 5, "source_updated_at": "2026-08-01",
        "salary_min": None, "salary_max": None,
        "first_seen_at": datetime(2026, 8, 9, 8, 0, tzinfo=_TAIWAN_TZ),
        "last_crawled_at": datetime(2026, 8, 9, 8, 0, tzinfo=_TAIWAN_TZ),
        "is_closed": False, "score": None, "recommend_reason": None, "skill_gap_note": None, "is_unliked": False,
    }
    row.update(overrides)
    return fake_db.insert("job_postings", row)


def test_list_scorable_jobs_only_includes_jobs_whose_company_has_background(fake_db):
    _insert_company(fake_db, "100", background="已回填的背景")
    _insert_company(fake_db, "200", background=None)
    _insert_job(fake_db, "1", "100")
    _insert_job(fake_db, "2", "200")

    jobs = job_search.list_scorable_jobs(fake_db)

    assert [j["job_id_104"] for j in jobs] == ["1"]


def test_list_scorable_jobs_returns_empty_when_no_company_has_background(fake_db):
    _insert_company(fake_db, "100", background=None)
    _insert_job(fake_db, "1", "100")

    assert job_search.list_scorable_jobs(fake_db) == []


# --- score_jobs／apply_scores（FR-37b～FR-37c）---


_PROFILE = {
    "resume": "五年後端工程師經驗", "years_of_experience": 5, "expected_salary_min": 60000,
    "expected_salary_max": 80000, "expectation": "想找後端或 AI 相關職缺",
}


def test_score_jobs_parses_batch_response_and_maps_to_job_id(fake_db):
    _insert_company(fake_db, "100")
    job = _insert_job(fake_db, "1", "100")
    jobs = fake_db.select("job_postings", where="id = %s", params=(job,))
    companies_by_id = {"100": {"company_name": "100 公司", "background": "已回填的背景", "industry": "軟體業"}}
    llm_client = MagicMock()
    llm_client.generate_text.return_value = "===JOB 1===\nSCORE: 88.5\nREASON: 職缺與履歷高度相符\nGAP: 缺乏雲端經驗"

    scores = job_search.score_jobs(llm_client, _PROFILE, jobs, companies_by_id)

    assert scores == {
        "1": {"score": 88.5, "recommend_reason": "職缺與履歷高度相符", "skill_gap_note": "缺乏雲端經驗"}
    }


def test_score_jobs_skips_job_when_llm_response_missing_that_block(fake_db):
    _insert_company(fake_db, "100")
    job1 = _insert_job(fake_db, "1", "100")
    job2 = _insert_job(fake_db, "2", "100")
    jobs = [
        j for j in fake_db.select("job_postings")
        if j["id"] in (job1, job2)
    ]
    companies_by_id = {"100": {"company_name": "100 公司", "background": "已回填的背景", "industry": "軟體業"}}
    llm_client = MagicMock()
    # 只回傳職缺 1 的區塊，職缺 2 完全沒出現在回應裡（模擬格式跑掉的情況）。
    llm_client.generate_text.return_value = "===JOB 1===\nSCORE: 70\nREASON: 尚可\nGAP: 無明顯落差"

    scores = job_search.score_jobs(llm_client, _PROFILE, jobs, companies_by_id)

    assert set(scores.keys()) == {"1"}


def test_score_jobs_batches_requests_when_exceeding_batch_size(fake_db):
    _insert_company(fake_db, "100")
    for i in range(1, 17):  # 16 筆，超過 _SCORING_BATCH_SIZE=15，應分兩批
        _insert_job(fake_db, str(i), "100")
    jobs = fake_db.select("job_postings")
    companies_by_id = {"100": {"company_name": "100 公司", "background": "已回填的背景", "industry": "軟體業"}}
    llm_client = MagicMock()
    llm_client.generate_text.return_value = "===JOB 1===\nSCORE: 50\nREASON: r\nGAP: g"

    job_search.score_jobs(llm_client, _PROFILE, jobs, companies_by_id)

    assert llm_client.generate_text.call_count == 2


def test_apply_scores_updates_matching_rows(fake_db):
    _insert_company(fake_db, "100")
    _insert_job(fake_db, "1", "100")
    scores = {"1": {"score": 92.0, "recommend_reason": "很符合", "skill_gap_note": "無明顯落差"}}

    updated = job_search.apply_scores(fake_db, scores)

    assert updated == 1
    row = fake_db.select("job_postings", where="job_id_104 = %s", params=("1",), fetch_one=True)
    assert row["score"] == 92.0
    assert row["recommend_reason"] == "很符合"
    assert row["skill_gap_note"] == "無明顯落差"


# --- build_ranked_jobs（FR-38a）---


def test_build_ranked_jobs_all_scope_excludes_unscored_unliked_and_closed(fake_db):
    _insert_company(fake_db, "100")
    _insert_job(fake_db, "1", "100", score=90.0)
    _insert_job(fake_db, "2", "100", score=None)  # 尚未評分
    _insert_job(fake_db, "3", "100", score=80.0, is_unliked=True)  # 已標記不喜歡
    _insert_job(fake_db, "4", "100", score=70.0, is_closed=True)  # 已關閉

    ranked = job_search.build_ranked_jobs(fake_db, "all")

    assert [j["job_id_104"] for j in ranked] == ["1"]
    assert ranked[0]["rank"] == 1


def test_build_ranked_jobs_sorted_by_score_descending(fake_db):
    _insert_company(fake_db, "100")
    _insert_job(fake_db, "1", "100", score=70.0)
    _insert_job(fake_db, "2", "100", score=95.0)
    _insert_job(fake_db, "3", "100", score=85.0)

    ranked = job_search.build_ranked_jobs(fake_db, "all")

    assert [j["job_id_104"] for j in ranked] == ["2", "3", "1"]
    assert [j["rank"] for j in ranked] == [1, 2, 3]


def test_build_ranked_jobs_limits_to_top_30(fake_db):
    _insert_company(fake_db, "100")
    for i in range(1, 35):
        _insert_job(fake_db, str(i), "100", score=float(i))

    ranked = job_search.build_ranked_jobs(fake_db, "all")

    assert len(ranked) == 30
    assert ranked[0]["job_id_104"] == "34"  # 分數最高


def test_build_ranked_jobs_new_this_week_scope_filters_by_first_seen_today(fake_db):
    _insert_company(fake_db, "100")
    today = date(2026, 8, 10)
    _insert_job(
        fake_db, "1", "100", score=90.0,
        first_seen_at=datetime(2026, 8, 10, 8, 0, tzinfo=_TAIWAN_TZ),
    )
    _insert_job(
        fake_db, "2", "100", score=95.0,
        first_seen_at=datetime(2026, 8, 3, 8, 0, tzinfo=_TAIWAN_TZ),
    )

    ranked = job_search.build_ranked_jobs(fake_db, "new_this_week", today=today)

    assert [j["job_id_104"] for j in ranked] == ["1"]


# --- build_job_recommendation_excel（FR-38b）---


def test_job_recommendation_excel_filename():
    assert job_search.job_recommendation_excel_filename(date(2026, 8, 9)) == "2026-08-09-104職缺推薦.xlsx"


def test_build_job_recommendation_excel_has_three_sheets_with_expected_headers():
    all_ranked = [
        {
            "rank": 1, "job_id_104": "1", "company_id_104": "100", "company_name": "A 公司",
            "region": "台北市", "industry": "軟體業", "title": "AI 工程師", "score": 90.0,
            "recommend_reason": "很符合", "skill_gap_note": "缺乏雲端經驗", "url": "https://www.104.com.tw/job/1",
        }
    ]
    new_ranked = []

    xlsx_bytes = job_search.build_job_recommendation_excel(all_ranked, new_ranked)

    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    assert workbook.sheetnames == ["所有職缺推薦", "最新職缺推薦", "技能缺口"]

    all_sheet = workbook["所有職缺推薦"]
    header = [cell.value for cell in next(all_sheet.iter_rows(min_row=1, max_row=1))]
    assert header == [
        "104公司ID", "公司全名", "地區", "產業類型", "職缺", "評分", "排名", "推薦原因", "連結", "是否喜歡",
    ]
    data_row = [cell.value for cell in next(all_sheet.iter_rows(min_row=2, max_row=2))]
    # openpyxl 讀回空字串儲存格時是 None（不是 ""），"是否喜歡" 欄位留給 Robin 手動標記。
    assert data_row == ["100", "A 公司", "台北市", "軟體業", "AI 工程師", 90.0, 1, "很符合", "https://www.104.com.tw/job/1", None]

    new_sheet = workbook["最新職缺推薦"]
    assert new_sheet.max_row == 1  # 只有表頭，沒有資料列

    gap_sheet = workbook["技能缺口"]
    gap_header = [cell.value for cell in next(gap_sheet.iter_rows(min_row=1, max_row=1))]
    assert gap_header == ["104職缺ID", "說明"]
    gap_row = [cell.value for cell in next(gap_sheet.iter_rows(min_row=2, max_row=2))]
    assert gap_row == ["1", "缺乏雲端經驗"]


def test_build_job_recommendation_excel_dedupes_skill_gap_rows_across_sheets():
    job = {
        "rank": 1, "job_id_104": "1", "company_id_104": "100", "company_name": "A 公司",
        "region": "台北市", "industry": "軟體業", "title": "AI 工程師", "score": 90.0,
        "recommend_reason": "很符合", "skill_gap_note": "缺乏雲端經驗", "url": "https://www.104.com.tw/job/1",
    }

    xlsx_bytes = job_search.build_job_recommendation_excel([job], [job])

    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    gap_sheet = workbook["技能缺口"]
    assert gap_sheet.max_row == 2  # 表頭 + 1 筆（同一職缺出現在兩張表也只記一次）


def test_send_job_recommendation_email_calls_client_with_expected_args():
    email_client = MagicMock()

    job_search.send_job_recommendation_email(email_client, "robin@gmail.com", date(2026, 8, 9), b"fake-xlsx-bytes")

    call_kwargs = email_client.send_text_with_attachment.call_args.kwargs
    assert call_kwargs["to"] == "robin@gmail.com"
    assert call_kwargs["subject"] == "2026-08-09 排程 - Robinson 104 職缺推薦"
    assert call_kwargs["body"] == "附件為本週整理的職缺推薦列表，以及技能缺口分析，請參閱！"
    assert call_kwargs["attachment_filename"] == "2026-08-09-104職缺推薦.xlsx"
    assert call_kwargs["attachment_bytes"] == b"fake-xlsx-bytes"


# --- parse_recommendation_excel／apply_job_preferences（FR-38e）---


def _build_test_recommendation_excel(rows):
    workbook = openpyxl.Workbook()
    for sheet_name in ("所有職缺推薦", "最新職缺推薦"):
        sheet = workbook.active if sheet_name == "所有職缺推薦" else workbook.create_sheet(sheet_name)
        sheet.title = sheet_name
        sheet.append(job_search._RECOMMENDATION_SHEET_HEADER)
        for row in rows:
            sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_parse_recommendation_excel_extracts_liked_flag_by_url():
    rows = [
        ["100", "A 公司", "台北市", "軟體業", "AI 工程師", 90.0, 1, "很符合", "https://www.104.com.tw/job/1", "1"],
        ["200", "B 公司", "新竹市", "硬體業", "後端工程師", 80.0, 2, "還不錯", "https://www.104.com.tw/job/2", ""],
    ]
    xlsx_bytes = _build_test_recommendation_excel(rows)

    entries = job_search.parse_recommendation_excel(xlsx_bytes)

    assert {"url": "https://www.104.com.tw/job/1", "is_unliked": True} in entries
    assert {"url": "https://www.104.com.tw/job/2", "is_unliked": False} in entries
    assert len(entries) == 2  # 兩張表同樣的連結去重後只留一筆


def test_apply_job_preferences_updates_matching_and_reports_not_found(fake_db):
    _insert_company(fake_db, "100")
    _insert_job(fake_db, "1", "100")
    entries = [
        {"url": "https://www.104.com.tw/job/1", "is_unliked": True},
        {"url": "https://www.104.com.tw/job/999", "is_unliked": False},
    ]

    result = job_search.apply_job_preferences(fake_db, entries)

    assert result == {"updated_count": 1, "not_found_urls": ["https://www.104.com.tw/job/999"]}
    row = fake_db.select("job_postings", where="job_id_104 = %s", params=("1",), fetch_one=True)
    assert row["is_unliked"] is True


# --- check_and_run_weekly_job_search 併入 FR-37/FR-38（FR-37d）---


def test_check_and_run_weekly_job_search_skips_scoring_when_llm_client_none(fake_db, monkeypatch):
    monkeypatch.setattr(job_search, "_polite_delay", lambda *a, **k: None)
    _seed_owner(fake_db)
    _insert_company(fake_db, "100", background="已回填的背景")
    email_client = MagicMock()
    telegram_client = MagicMock()

    job_search.check_and_run_weekly_job_search(
        fake_db, _FakeJob104Client(pages={}, details={}), email_client, "robin@gmail.com", telegram_client,
        llm_client=None, now=_MONDAY_8AM,
    )

    # llm_client 未提供時，評分/推薦信整段跳過，但不影響公司背景流程本身已跑過（此案例沒有新公司）。
    assert email_client.send_text_with_attachment.call_count == 0


def test_check_and_run_weekly_job_search_runs_scoring_and_sends_recommendation_email(fake_db, monkeypatch):
    monkeypatch.setattr(job_search, "_polite_delay", lambda *a, **k: None)
    owner_id = _seed_owner(fake_db)
    fake_db.update(
        "users",
        {
            "job_resume": "五年後端經驗", "job_expectation": "想找後端職缺",
            "years_of_experience": 5, "expected_salary_min": 60000, "expected_salary_max": 80000,
        },
        where="id = %s", params=(owner_id,),
    )
    _insert_company(fake_db, "100", background="已回填的背景")
    _insert_job(fake_db, "1", "100")  # 已存在資料庫、背景已回填，屬於這次評分範圍
    client = _FakeJob104Client(pages={}, details={})
    email_client = MagicMock()
    telegram_client = MagicMock()
    llm_client = MagicMock()
    llm_client.generate_text.return_value = "===JOB 1===\nSCORE: 90\nREASON: 很符合\nGAP: 無明顯落差"

    job_search.check_and_run_weekly_job_search(
        fake_db, client, email_client, "robin@gmail.com", telegram_client, llm_client=llm_client, now=_MONDAY_8AM
    )

    row = fake_db.select("job_postings", where="job_id_104 = %s", params=("1",), fetch_one=True)
    assert row["score"] == 90.0
    email_client.send_text_with_attachment.assert_called_once()
    call_kwargs = email_client.send_text_with_attachment.call_args.kwargs
    assert call_kwargs["attachment_filename"] == "2026-08-10-104職缺推薦.xlsx"
    telegram_client.send_text.assert_called_once_with(
        chat_id=8263904025, text=job_search.RECOMMENDATION_EMAIL_SENT_NOTIFICATION_TEXT
    )


def test_check_and_run_weekly_job_search_no_scorable_jobs_skips_recommendation_email(fake_db, monkeypatch):
    monkeypatch.setattr(job_search, "_polite_delay", lambda *a, **k: None)
    _seed_owner(fake_db)
    client = _FakeJob104Client(pages={}, details={})
    email_client = MagicMock()
    telegram_client = MagicMock()
    llm_client = MagicMock()

    job_search.check_and_run_weekly_job_search(
        fake_db, client, email_client, "robin@gmail.com", telegram_client, llm_client=llm_client, now=_MONDAY_8AM
    )

    llm_client.generate_text.assert_not_called()
    email_client.send_text_with_attachment.assert_not_called()
