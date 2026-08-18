"""求職模組商業邏輯（Step 4.1，見 docs/specs/robinson/SPEC.md FR-33～FR-36，ADR-24）。

僅 Robin 一人可用（`job_search` 開關於 Step 4.1 改為 `owner_only=True`，見 ADR-24 決策 2），
所以本模組不像 `finance.py`／`body.py` 需要處理多使用者並行的複雜度。涵蓋 FR-33（搜尋條件）、
FR-34（兩階段爬蟲＋ETL 去重，呼叫 `submodules/job104` 拿資料）、FR-36（履歷／期望工作敘述／
結構化年資與期望薪資）；FR-35（公司背景 Email 協作）見本檔案下方區塊。
"""
import csv
import io
import logging
import random
import re
import time
import uuid
from datetime import date, datetime
from zoneinfo import ZoneInfo

import openpyxl

from src.bot import toggles
from submodules.cloudsql.client import CloudSQLClient

_logger = logging.getLogger(__name__)

_TAIWAN_TZ = ZoneInfo("Asia/Taipei")

_FEATURE_KEY = "job_search"

# FR-34b：每週僅執行一次，固定台灣時間週一 08:00（Robin 2026-08-09 確認，比照 YouTube 模組
# 每週四 08:00 的既有慣例，見 youtube.py `_PUSH_WEEKDAY_THURSDAY`）。
_WEEKLY_CRAWL_WEEKDAY = 0  # Python datetime.weekday()：Monday=0
_WEEKLY_CRAWL_HOUR = 8

# FR-36：履歷與期望工作敘述皆為「3500 字以內」，兩者上限相同。
_MAX_TEXT_LENGTH = 3500

# FR-36：年資（`years_of_experience`）合理範圍，0～60 年已足夠涵蓋所有實際情境，避免使用者
# 誤植成月份數字（例如打「36」代表 3 年又半，這裡不猜測，一律要求使用者輸入年為單位的數字）。
_MIN_YEARS_OF_EXPERIENCE = 0
_MAX_YEARS_OF_EXPERIENCE = 60

# FR-34c：列表分頁／詳情頁請求之間強制 2～4 秒隨機延遲，嚴禁併發多執行緒請求。
_MIN_DELAY_SECONDS = 2
_MAX_DELAY_SECONDS = 4

# 安全防呆（非 spec 硬性規定）：避免 104 因未知原因對某組條件一直回傳非空清單時無限迴圈，
# 20 頁（每頁通常 20 筆，約 400 筆）已遠超過個人使用情境下單一搜尋條件實際會出現的職缺量。
_MAX_PAGES_PER_CRITERIA = 20


def is_text_length_valid(text: str) -> bool:
    """FR-36：履歷／期望工作敘述皆不可超過 3500 字。"""
    return len(text) <= _MAX_TEXT_LENGTH


def is_years_of_experience_reasonable(years: float) -> bool:
    """FR-36：年資合理範圍檢查，0～60 年。"""
    return _MIN_YEARS_OF_EXPERIENCE <= years <= _MAX_YEARS_OF_EXPERIENCE


def save_search_criteria(
    db: CloudSQLClient,
    user_id: int,
    keyword: str,
    region: str | None,
    salary_min: int | None,
    salary_max: int | None,
) -> int:
    """新增一組求職搜尋條件（FR-33），回傳新建列的 id。

    ADR-24 決策 3：允許同時存多組條件，不比照記帳預算/證照目標「一人一份、重新設定即覆蓋」的
    既有慣例，這裡固定用 INSERT 新增一筆，不會動到既有的其他組條件。

    2026-08-09 依 Robin 指示移除產業篩選（`industry`）——104 實測驗證後發現這個維度不值得繼續
    猜測參數名稱。`job_search_criteria.industry` 欄位保留在資料庫（不做 migration 刪除，避免
    非必要的破壞性操作），只是往後一律不再寫入、也不再從對話流程收集。
    """
    return db.insert(
        "job_search_criteria",
        {
            "user_id": user_id,
            "keyword": keyword,
            "region": region,
            "salary_min": salary_min,
            "salary_max": salary_max,
        },
    )


def save_profile(
    db: CloudSQLClient,
    user_id: int,
    resume: str,
    expectation: str,
    years_of_experience: float,
    expected_salary_min: int,
    expected_salary_max: int,
) -> None:
    """寫入使用者履歷／期望工作敘述／結構化年資與期望薪資（FR-36）。

    這五個欄位一次一起寫入（收集流程走到最後一步才呼叫，見 `src/bot/commands.py`
    `handle_job_search_salary_max_step`），不提供只更新部分欄位的介面——FR-36 的設計就是
    一輪完整對話收集齊全，重新執行整個 `/set_job_search` 流程即可覆蓋更新。
    """
    db.update(
        "users",
        {
            "job_resume": resume,
            "job_expectation": expectation,
            "years_of_experience": years_of_experience,
            "expected_salary_min": expected_salary_min,
            "expected_salary_max": expected_salary_max,
        },
        where="id = %s",
        params=(user_id,),
    )


# --- FR-34：104 職缺爬蟲（兩階段架構＋ETL 去重）---


def _now() -> datetime:
    """回傳現在的台灣時間；獨立成函式方便測試用 monkeypatch 固定時間點。"""
    return datetime.now(_TAIWAN_TZ)


def list_search_criteria(db: CloudSQLClient, user_id: int) -> list[dict]:
    """查詢使用者目前設定的所有求職搜尋條件（FR-33），供 FR-34 週排程逐組送出查詢。"""
    return db.select("job_search_criteria", where="user_id = %s", params=(user_id,))


def _polite_delay(sleep_func, random_func) -> None:
    """FR-34c：每次對 104 發出請求後呼叫一次，強制 2～4 秒隨機延遲才能發下一次請求。"""
    sleep_func(random_func(_MIN_DELAY_SECONDS, _MAX_DELAY_SECONDS))


def upsert_company(
    db: CloudSQLClient, company_id_104: str, company_name: str, region: str, industry: str | None = None
) -> tuple[int, bool]:
    """FR-35a：找出這批職缺所屬公司是否已存在資料庫（以 104 公司 ID 判斷），不存在才新增
    （`background` 留空，代表尚待 Robin 人工回填）；已存在則完全不覆蓋既有欄位（避免蓋掉已經
    回填好的背景資料或更新過的公司資訊）。回傳 `(job_companies.id, 是否為新公司)`。
    """
    existing = db.select("job_companies", where="company_id_104 = %s", params=(company_id_104,), fetch_one=True)
    if existing is not None:
        return existing["id"], False

    new_id = db.insert(
        "job_companies",
        {
            "company_id_104": company_id_104,
            "company_name": company_name,
            "region": region,
            "industry": industry,
            "background": None,
        },
    )
    return new_id, True


def upsert_job_posting(db: CloudSQLClient, job: dict, detail: dict, now: datetime) -> bool:
    """FR-34d ETL 去重：以 `job_id_104` 為鍵值，已存在就 `UPDATE`（薪資/內容等可能隨時間變動的
    欄位＋`last_crawled_at`），不存在才 `INSERT` 新增一筆（`first_seen_at` 只在新增當下寫入一次，
    供 FR-38a「本週新職缺排名」判斷依據，之後每次重新爬到都不會再變動）。回傳「是否為新職缺」。

    薪資（`salary_min`／`salary_max`）目前不寫入——104 列表/詳情 API 的薪資描述格式尚未經過
    真實流量驗證（見 `submodules/job104/client.py` 模組 docstring），與其現在用不確定的規則
    猜測解析成數字區間，不如先讓這兩欄維持 `NULL`，待實測結果出爐後再補上解析邏輯，符合 ADR-24
    決策 4「先保守假設、之後視實測結果調整」的一貫做法。

    應徵人數（`applicant_count`）取自 `job`（列表階段的 `search_list()` 結果），不是
    `detail`（詳情頁 API 已確認沒有這個欄位，2026-08-09 實測驗證，見 `submodules/job104/
    client.py` 模組 docstring）。

    是否已關閉（`is_closed`）同樣取自 `job`（`search_list()` 依 `jobSwitch` 欄位自動判斷，
    2026-08-09 實測驗證確認可行，見 ADR-26 決策 5）；已存在的職缺重新爬到時一併更新這個
    欄位，職缺重新開放/關閉的狀態變化才能反映到資料庫，不會卡在第一次爬到時的舊狀態。
    """
    existing = db.select("job_postings", where="job_id_104 = %s", params=(job["job_id"],), fetch_one=True)
    fields = {
        "company_id_104": job["company_id"],
        "title": job["title"],
        "region": job["region"],
        "url": job["url"],
        "content": detail.get("content"),
        "required_years_experience": detail.get("required_years_experience"),
        "applicant_count": job.get("applicant_count"),
        "source_updated_at": detail.get("source_updated_at"),
        "is_closed": job.get("is_closed", False),
        "last_crawled_at": now,
    }
    if existing is not None:
        if existing.get("is_closed_manual_override"):
            fields.pop("is_closed")
        db.update("job_postings", fields, where="job_id_104 = %s", params=(job["job_id"],))
        return False

    db.insert(
        "job_postings",
        {**fields, "job_id_104": job["job_id"], "first_seen_at": now, "is_closed_manual_override": False},
    )
    return True


def crawl_and_upsert_jobs(
    db: CloudSQLClient,
    job104_client,
    user_id: int,
    sleep_func=time.sleep,
    random_func=random.uniform,
    now: datetime | None = None,
) -> dict:
    """FR-34：對這位使用者目前設定的每一組搜尋條件（FR-33），依 FR-34a 兩階段架構爬取職缺——
    先呼叫列表 API 依條件取得摘要清單，再對清單內每一筆職缺 ID 個別呼叫詳情頁補齊完整內容。
    FR-34c 的 2～4 秒隨機延遲套用在**每一次**對 104 發出的請求之後（不論列表分頁或詳情頁），
    確保連續兩次請求之間一定間隔至少 2 秒；FR-34c 嚴禁併發，本函式全程單執行緒依序呼叫，
    天生滿足這個限制，不需要額外的鎖或旗標。FR-34d 的 ETL 去重見 `upsert_job_posting()`。

    列表 API 回傳空清單（含翻到最後一頁之後）視為這組條件已經爬完，換下一組條件；單組條件
    最多翻 `_MAX_PAGES_PER_CRITERIA` 頁（安全防呆，非 spec 硬性規定，見該常數說明）。

    地區篩選（`criteria["region"]`）2026-08-09 起改為**呼叫端自行做子字串比對篩選**，不送給
    104 API——104 的 `area` 參數要傳它自己的地區數字代碼，不是使用者輸入的地區文字，沒有可靠
    對照表（見 `submodules/job104/client.py` 模組 docstring）。分頁停止判斷仍然依據**未篩選**
    的原始清單是否為空（`jobs`），不是篩選後的 `matching_jobs`——避免某一頁剛好篩不到符合地區
    的職缺，就誤判成「這組條件已經爬完」而提早停止翻頁。

    回傳這次爬蟲的統計摘要：`{"new_company_ids": [...], "new_job_count": int,
    "updated_job_count": int}`；`new_company_ids` 供 FR-35a 判斷這批職缺所屬公司是否需要
    走 Email/CSV/Drive 協作流程（見 `build_new_companies_csv()`）。
    """
    now = now or _now()
    criteria_list = list_search_criteria(db, user_id)

    new_company_ids: list[str] = []
    new_job_count = 0
    updated_job_count = 0

    for criteria in criteria_list:
        region_filter = criteria.get("region")
        page = 1
        while page <= _MAX_PAGES_PER_CRITERIA:
            jobs = job104_client.search_list(
                criteria["keyword"],
                salary_min=criteria.get("salary_min"),
                salary_max=criteria.get("salary_max"),
                page=page,
            )
            _polite_delay(sleep_func, random_func)
            if not jobs:
                break

            matching_jobs = [j for j in jobs if not region_filter or region_filter in (j.get("region") or "")]
            for job in matching_jobs:
                _, is_new_company = upsert_company(db, job["company_id"], job["company_name"], job["region"])
                if is_new_company and job["company_id"] not in new_company_ids:
                    new_company_ids.append(job["company_id"])

                detail = job104_client.fetch_job_detail(job["job_slug"])
                _polite_delay(sleep_func, random_func)

                if upsert_job_posting(db, job, detail, now):
                    new_job_count += 1
                else:
                    updated_job_count += 1

            page += 1

    return {
        "new_company_ids": new_company_ids,
        "new_job_count": new_job_count,
        "updated_job_count": updated_job_count,
    }


# --- FR-35：公司背景 Email／CSV／Drive 人力協作機制（見 ADR-24 決策 1）---

_COMPANY_CSV_HEADER = ["104公司ID", "公司全名", "地區", "產業類型", "背景"]
_COMPANY_CSV_FILENAME_SUFFIX = "-104職缺公司.csv"

# FR-35b：固定信件標題／內文文案（Robin 已核准的措辭，比照 templates.py 靜態文案不經 LLM 生成）。
EMAIL_SUBJECT_TEMPLATE = "{date} 排程 - Robinson 104 職缺公司列表"
EMAIL_BODY_TEXT = "附件為本週爬到的最新公司列表，請參閱！"

# FR-35c：寄信成功後私訊 Robin 的固定文案（FR-19h 決策執行狀態閉環回饋）。
EMAIL_SENT_NOTIFICATION_TEXT = "已經寄送本週最新的104職缺公司信件給您了～"


def company_csv_filename(target_date: date) -> str:
    """FR-35b：組出公司背景 CSV 的固定檔名格式 `{YYYY-MM-DD}-104職缺公司.csv`。"""
    return f"{target_date.isoformat()}{_COMPANY_CSV_FILENAME_SUFFIX}"


def get_companies_by_ids(db: CloudSQLClient, company_ids: list[str]) -> list[dict]:
    """依 104 公司 ID 清單查出對應的 `job_companies` 資料列，供 FR-35b 組 CSV 使用。

    逐一查詢再彙整，不做一次性的 `IN` 查詢——個人使用情境下單週新公司數量不會太多，不需要
    為此另外設計 `FakeCloudSQLClient`／正式 `CloudSQLClient` 都要支援的批次查詢介面。
    """
    rows = []
    for company_id in company_ids:
        row = db.select("job_companies", where="company_id_104 = %s", params=(company_id,), fetch_one=True)
        if row is not None:
            rows.append(row)
    return rows


def build_new_companies_csv(companies: list[dict]) -> str:
    """FR-35b：組出待 Robin 查詢回填的新公司 CSV 內容（欄位：104公司ID／公司全名／地區／
    產業類型／背景）。用標準函式庫 `csv` module，不額外安裝第三方套件（比照
    `submodules/newsfeed` 用標準函式庫解析 RSS 而不裝 `feedparser` 的既有慣例，見 ADR-24
    後果）。「背景」欄位固定留空字串，等 Robin 查完手動填入。
    """
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(_COMPANY_CSV_HEADER)
    for company in companies:
        writer.writerow(
            [
                company["company_id_104"],
                company["company_name"],
                company.get("region") or "",
                company.get("industry") or "",
                "",
            ]
        )
    return buffer.getvalue()


def send_new_companies_email(email_client, to: str, target_date: date, csv_text: str) -> None:
    """FR-35b：把公司背景 CSV 當附件寄給 Robin 自己（`GMAIL_USER` 自寄自收）。"""
    email_client.send_text_with_attachment(
        to=to,
        subject=EMAIL_SUBJECT_TEMPLATE.format(date=target_date.isoformat()),
        body=EMAIL_BODY_TEXT,
        attachment_filename=company_csv_filename(target_date),
        attachment_bytes=csv_text.encode("utf-8-sig"),  # BOM：Excel 開啟中文 CSV 不亂碼
    )


def parse_companies_csv(csv_text: str) -> list[dict]:
    """FR-35e：解析 Robin 回填好的公司背景 CSV，回傳 `[{"company_id_104", "background"}, ...]`。

    「104公司ID」或「背景」任一欄位為空的列直接跳過（分別代表格式異常、或 Robin 還沒查完
    這家公司），不會把空字串誤當成「已確認查無資料」寫回資料庫覆蓋掉未來可能補上的內容。
    """
    reader = csv.DictReader(io.StringIO(csv_text))
    entries = []
    for row in reader:
        company_id = (row.get("104公司ID") or "").strip()
        background = (row.get("背景") or "").strip()
        if not company_id or not background:
            continue
        entries.append({"company_id_104": company_id, "background": background})
    return entries


def apply_company_backgrounds(db: CloudSQLClient, entries: list[dict]) -> dict:
    """FR-35e：把解析出的背景逐筆 `UPDATE` 回 `job_companies`（以 104 公司 ID 比對）。

    回傳 `{"updated_count": int, "not_found_ids": [...]}`；比對不到對應公司的 ID 一律列出來，
    不可靜默略過（比照 FR-38e「找不到對應職缺就列出來提醒人工處理」的一貫做法）。
    """
    updated_count = 0
    not_found_ids: list[str] = []
    for entry in entries:
        affected = db.update(
            "job_companies",
            {"background": entry["background"]},
            where="company_id_104 = %s",
            params=(entry["company_id_104"],),
        )
        if affected:
            updated_count += 1
        else:
            not_found_ids.append(entry["company_id_104"])
    return {"updated_count": updated_count, "not_found_ids": not_found_ids}


# --- FR-37：Gemini 批次契合度評分（見 ADR-26 決策 2～4）---

# FR-37c：職缺數量過多超出單次 Prompt 負荷時分批送出，比照 youtube.py 批次評分的做法，每批
# 獨立呼叫一次 Gemini；15 筆是保守估計值（每筆職缺內容含完整 content/welfare 文字，單批太大
# 容易讓 LLM 輸出格式跑掉或超出 token 上限）。
_SCORING_BATCH_SIZE = 15

_SCORING_PROMPT_TEMPLATE = (
    "你是一位職涯顧問，請根據使用者的履歷、年資、期望薪資，針對下面每一筆待評分職缺，計算一個"
    "0～100 的整體契合度分數（分數越高代表越推薦，需綜合考量職缺內容、年資落差、期望薪資落差、"
    "應徵熱門程度、更新時間新舊），並各自給出簡短的推薦原因與技能缺口說明。若某項資訊標示"
    "「未提供」，評分時直接略過該項比對維度，不要用未提供的資訊臆測。\n\n"
    "【使用者履歷】\n{resume}\n\n"
    "【使用者年資】{years_of_experience} 年\n"
    "【使用者期望薪資】{expected_salary_min}～{expected_salary_max} 元\n"
    "【使用者期望工作內容】\n{expectation}\n\n"
    "【待評分職缺清單】\n{jobs_text}\n\n"
    "請針對每一筆職缺，嚴格依照下面格式各自輸出一個區塊，區塊之間不要有其他文字：\n"
    "===JOB <職缺編號>===\n"
    "SCORE: <0~100 的分數，可為小數>\n"
    "REASON: <推薦原因，50 字以內，限一行>\n"
    "GAP: <技能缺口說明，100 字以內，限一行，沒有明顯缺口可寫「無明顯落差」>"
)

_SCORE_BLOCK_PATTERN = re.compile(r"===\s*JOB\s+(\d+)\s*===")
_SCORE_FIELD_PATTERN = re.compile(r"SCORE:\s*(\d+(?:\.\d+)?)")
_REASON_FIELD_PATTERN = re.compile(r"REASON:\s*(.+)")
_GAP_FIELD_PATTERN = re.compile(r"GAP:\s*(.+)")


def list_scorable_jobs(db: CloudSQLClient) -> list[dict]:
    """FR-37a：撈出「所屬公司背景資料已回填完成」的職缺，供每週批次評分使用；背景仍空白的職缺
    這次跳過，之後背景補齊會在下次排程自然被納入，不會卡住整批評分。

    `CloudSQLClient` 介面沒有 JOIN 查詢（`FakeCloudSQLClient` 測試替身也不支援），改成先查出
    背景已回填的公司 ID 集合，再用集合過濾職缺，於 Python 端完成等效的過濾邏輯。
    """
    companies_with_background = db.select("job_companies", where="background IS NOT NULL")
    company_ids = {c["company_id_104"] for c in companies_with_background}
    if not company_ids:
        return []
    return [job for job in db.select("job_postings") if job["company_id_104"] in company_ids]


def _format_job_for_prompt(index: int, job: dict, company: dict) -> str:
    required_years = job.get("required_years_experience")
    salary_min = job.get("salary_min")
    salary_max = job.get("salary_max")
    salary_text = f"{salary_min}～{salary_max} 元" if salary_min is not None or salary_max is not None else "未提供"
    return (
        f"{index}. 職缺：{job['title']}\n"
        f"   公司：{company.get('company_name', '未知')}（{company.get('industry') or '未提供'}）\n"
        f"   地區：{job.get('region') or '未提供'}\n"
        f"   要求年資：{required_years if required_years is not None else '未提供'} 年\n"
        f"   薪資範圍：{salary_text}\n"
        f"   應徵人數：{job.get('applicant_count') if job.get('applicant_count') is not None else '未提供'}　"
        f"更新時間：{job.get('source_updated_at') or '未提供'}\n"
        f"   公司背景：{company.get('background') or '未提供'}\n"
        f"   職缺內容：{job.get('content') or '未提供'}"
    )


def _build_scoring_prompt(profile: dict, jobs: list[dict], companies_by_id: dict) -> str:
    jobs_text = "\n".join(
        _format_job_for_prompt(index, job, companies_by_id.get(job["company_id_104"], {}))
        for index, job in enumerate(jobs, start=1)
    )
    return _SCORING_PROMPT_TEMPLATE.format(
        resume=profile.get("resume") or "未提供",
        years_of_experience=profile.get("years_of_experience"),
        expected_salary_min=profile.get("expected_salary_min"),
        expected_salary_max=profile.get("expected_salary_max"),
        expectation=profile.get("expectation") or "未提供",
        jobs_text=jobs_text,
    )


def _parse_scoring_response(raw: str, count: int) -> dict[int, dict]:
    """解析 Gemini 依 `_SCORING_PROMPT_TEMPLATE` 格式輸出的批次評分回應，回傳
    `{批次內編號: {"score", "reason", "gap"}}`；單一區塊缺少 SCORE 欄位（格式跑掉）直接跳過這筆，
    不強行湊資料，也不讓整批解析失敗（見 `score_jobs()` 的略過邏輯）。
    """
    segments = _SCORE_BLOCK_PATTERN.split(raw)
    # re.split 搭配 capture group 回傳：[分隔前綴, 編號1, 區塊1, 編號2, 區塊2, ...]
    results: dict[int, dict] = {}
    for i in range(1, len(segments), 2):
        try:
            index = int(segments[i])
        except ValueError:
            continue
        if not (1 <= index <= count):
            continue
        block = segments[i + 1]
        score_match = _SCORE_FIELD_PATTERN.search(block)
        if not score_match:
            continue
        reason_match = _REASON_FIELD_PATTERN.search(block)
        gap_match = _GAP_FIELD_PATTERN.search(block)
        results[index] = {
            "score": float(score_match.group(1)),
            "reason": reason_match.group(1).strip() if reason_match else "",
            "gap": gap_match.group(1).strip() if gap_match else "",
        }
    return results


def score_jobs(llm_client, profile: dict, jobs: list[dict], companies_by_id: dict) -> dict[str, dict]:
    """FR-37：把 `list_scorable_jobs()` 範圍內的職缺整批（而非逐筆）交給 Gemini 計算契合度分數、
    推薦原因、技能缺口說明（FR-37b～FR-37c，比照 youtube.py `score_candidates_for_topic()` 的
    批次評分模式，見模組上方 `_SCORING_BATCH_SIZE` 說明）。單一職缺解析不出有效分數時（LLM 輸出
    格式跑掉），這筆職缺這次直接跳過（`score` 維持 `NULL`），下週排程會重新嘗試，不強行湊資料也
    不讓整個評分流程卡住。

    回傳 `{job_id_104: {"score", "recommend_reason", "skill_gap_note"}}`，供 `apply_scores()`
    批次寫回資料庫。
    """
    results: dict[str, dict] = {}
    for start in range(0, len(jobs), _SCORING_BATCH_SIZE):
        batch = jobs[start : start + _SCORING_BATCH_SIZE]
        prompt = _build_scoring_prompt(profile, batch, companies_by_id)
        raw = llm_client.generate_text(prompt)
        parsed = _parse_scoring_response(raw, len(batch))
        for index, job in enumerate(batch, start=1):
            if index not in parsed:
                _logger.warning("職缺評分 LLM 回應解析失敗，略過這筆（job_id_104=%s）", job["job_id_104"])
                continue
            entry = parsed[index]
            results[job["job_id_104"]] = {
                "score": entry["score"],
                "recommend_reason": entry["reason"],
                "skill_gap_note": entry["gap"],
            }
    return results


def apply_scores(db: CloudSQLClient, scores: dict) -> int:
    """把 `score_jobs()` 的結果批次 `UPDATE` 回 `job_postings`，回傳實際更新筆數。"""
    updated = 0
    for job_id_104, fields in scores.items():
        affected = db.update(
            "job_postings",
            {
                "score": fields["score"],
                "recommend_reason": fields["recommend_reason"],
                "skill_gap_note": fields["skill_gap_note"],
            },
            where="job_id_104 = %s",
            params=(job_id_104,),
        )
        updated += affected
    return updated


# --- FR-38a：雙重排名（全庫／本週新職缺，動態計算不持久化，見 migration 0058 設計理由）---

_RANKING_LIMIT = 30


def _companies_by_id(db: CloudSQLClient) -> dict[str, dict]:
    return {c["company_id_104"]: c for c in db.select("job_companies")}


def _to_taiwan_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.astimezone(_TAIWAN_TZ).date()
    return value


def build_ranked_jobs(
    db: CloudSQLClient,
    scope: str,
    companies_by_id: dict | None = None,
    today: date | None = None,
    limit: int = _RANKING_LIMIT,
) -> list[dict]:
    """FR-38a：依 `score` 由高到低排出前 `limit` 名職缺（預設 30）。`scope="all"` 為全庫排名，
    `scope="new_this_week"` 僅計入 `first_seen_at`（換算台灣時間）等於 `today` 的職缺，也就是
    本次排程新爬到的職缺。排除 `score IS NULL`（尚未評分）、`is_unliked = TRUE`、
    `is_closed = TRUE` 的職缺；`rank` 在這裡動態計算，不持久化存進資料庫。
    """
    companies_by_id = companies_by_id if companies_by_id is not None else _companies_by_id(db)
    today = today or _now().date()

    eligible = [
        job
        for job in db.select("job_postings")
        if job.get("score") is not None and not job.get("is_unliked") and not job.get("is_closed")
    ]
    if scope == "new_this_week":
        eligible = [job for job in eligible if _to_taiwan_date(job.get("first_seen_at")) == today]
    eligible.sort(key=lambda job: float(job["score"]), reverse=True)

    ranked = []
    for rank, job in enumerate(eligible[:limit], start=1):
        company = companies_by_id.get(job["company_id_104"], {})
        ranked.append(
            {
                "rank": rank,
                "job_id_104": job["job_id_104"],
                "company_id_104": job["company_id_104"],
                "company_name": company.get("company_name", ""),
                "region": job.get("region") or company.get("region") or "",
                "industry": company.get("industry") or "",
                "title": job["title"],
                "score": job["score"],
                "recommend_reason": job.get("recommend_reason") or "",
                "skill_gap_note": job.get("skill_gap_note") or "",
                "url": job["url"],
            }
        )
    return ranked


# --- FR-38b：職缺推薦 Excel（三張工作表）---

_RECOMMENDATION_EXCEL_FILENAME_SUFFIX = "-104職缺推薦.xlsx"
# FR-39a（2026-08-09 追加，Step 4.3，見 ADR-27）：新增「104職缺ID」欄位，Robin 打「ID=XXX 職缺
# 已應徵」這類語句時直接從這裡抄，不用另外跳去「技能缺口」工作表查。
_RECOMMENDATION_SHEET_HEADER = [
    "104職缺ID", "104公司ID", "公司全名", "地區", "產業類型", "職缺", "評分", "排名", "推薦原因", "連結", "是否喜歡",
]
_SKILL_GAP_SHEET_HEADER = ["104職缺ID", "說明"]


def job_recommendation_excel_filename(target_date: date) -> str:
    """FR-38b：組出職缺推薦 Excel 的固定檔名格式 `{YYYY-MM-DD}-104職缺推薦.xlsx`。"""
    return f"{target_date.isoformat()}{_RECOMMENDATION_EXCEL_FILENAME_SUFFIX}"


def _write_recommendation_rows(sheet, ranked_jobs: list[dict]) -> None:
    sheet.append(_RECOMMENDATION_SHEET_HEADER)
    for job in ranked_jobs:
        sheet.append(
            [
                job["job_id_104"], job["company_id_104"], job["company_name"], job["region"], job["industry"],
                job["title"], float(job["score"]), job["rank"], job["recommend_reason"], job["url"], "",
            ]
        )


def _write_skill_gap_rows(sheet, all_ranked: list[dict], new_ranked: list[dict]) -> None:
    sheet.append(_SKILL_GAP_SHEET_HEADER)
    seen: set[str] = set()
    for job in [*all_ranked, *new_ranked]:
        if job["job_id_104"] in seen:
            continue
        seen.add(job["job_id_104"])
        sheet.append([job["job_id_104"], job["skill_gap_note"]])


def build_job_recommendation_excel(all_ranked: list[dict], new_ranked: list[dict]) -> bytes:
    """FR-38b：組出職缺推薦 Excel（三張工作表：所有職缺推薦／最新職缺推薦／技能缺口），用
    `openpyxl`（全專案第一次需要真的讀寫 .xlsx，2026-08-09 新增依賴，見 `requirements.txt`）。

    「是否喜歡」欄位固定留空字串讓 Robin 標記（FR-38d：填 1 代表不喜歡）；「是否關閉」已於
    Step 4.1 確認可用 `job_postings.is_closed` 自動判斷（見 FR-38b 2026-08-09 更新），這裡不
    出現這欄。
    """
    workbook = openpyxl.Workbook()
    all_sheet = workbook.active
    all_sheet.title = "所有職缺推薦"
    _write_recommendation_rows(all_sheet, all_ranked)

    new_sheet = workbook.create_sheet("最新職缺推薦")
    _write_recommendation_rows(new_sheet, new_ranked)

    gap_sheet = workbook.create_sheet("技能缺口")
    _write_skill_gap_rows(gap_sheet, all_ranked, new_ranked)

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# FR-38b：固定信件標題／內文文案（Robin 已核准的措辭，比照 EMAIL_SUBJECT_TEMPLATE／EMAIL_BODY_TEXT）。
RECOMMENDATION_EMAIL_SUBJECT_TEMPLATE = "{date} 排程 - Robinson 104 職缺推薦"
RECOMMENDATION_EMAIL_BODY_TEXT = "附件為本週整理的職缺推薦列表，以及技能缺口分析，請參閱！"

# FR-38c：寄信成功後私訊 Robin 的固定文案。
RECOMMENDATION_EMAIL_SENT_NOTIFICATION_TEXT = "已寄送本週 104 職缺推薦檔案給您～"


def send_job_recommendation_email(email_client, to: str, target_date: date, xlsx_bytes: bytes) -> None:
    """FR-38b：把職缺推薦 Excel 當附件寄給 Robin 自己（`GMAIL_USER` 自寄自收）。"""
    email_client.send_text_with_attachment(
        to=to,
        subject=RECOMMENDATION_EMAIL_SUBJECT_TEMPLATE.format(date=target_date.isoformat()),
        body=RECOMMENDATION_EMAIL_BODY_TEXT,
        attachment_filename=job_recommendation_excel_filename(target_date),
        attachment_bytes=xlsx_bytes,
    )


# --- FR-38e：Robin 回填「是否喜歡」後回寫資料庫 ---


def parse_recommendation_excel(xlsx_bytes: bytes) -> list[dict]:
    """FR-38e：解析 Robin 回填好「是否喜歡」欄位的推薦 Excel，回傳 `[{"url", "is_unliked"}, ...]`。

    以「連結」（職缺 URL）為比對鍵值——天然唯一，不需要額外的比對欄位。「所有職缺推薦」／
    「最新職缺推薦」兩張工作表都要讀，同一個連結可能同時出現在兩張表，用 dict 依連結去重
    （後讀到的覆蓋前面的；實務上同一職缺兩張表的標記內容應該一致，不特別處理衝突）。
    """
    workbook = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    entries: dict[str, bool] = {}
    for sheet_name in ("所有職缺推薦", "最新職缺推薦"):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        if "連結" not in header or "是否喜歡" not in header:
            continue
        url_index = header.index("連結")
        liked_index = header.index("是否喜歡")
        for row in sheet.iter_rows(min_row=2, values_only=True):
            url = row[url_index]
            if not url:
                continue
            entries[url] = str(row[liked_index]).strip() == "1"
    return [{"url": url, "is_unliked": is_unliked} for url, is_unliked in entries.items()]


def apply_job_preferences(db: CloudSQLClient, entries: list[dict]) -> dict:
    """FR-38e：把解析出的「是否喜歡」標記逐筆 `UPDATE` 回 `job_postings`（以 url 比對）。

    回傳 `{"updated_count": int, "not_found_urls": [...]}`；比對不到對應職缺的一律列出來提醒
    人工處理，不可靜默略過（比照 FR-35e）。
    """
    updated_count = 0
    not_found_urls: list[str] = []
    for entry in entries:
        affected = db.update(
            "job_postings", {"is_unliked": entry["is_unliked"]}, where="url = %s", params=(entry["url"],)
        )
        if affected:
            updated_count += 1
        else:
            not_found_urls.append(entry["url"])
    return {"updated_count": updated_count, "not_found_urls": not_found_urls}


# --- 每週排程整合入口（FR-34b、FR-35a～FR-35c、FR-37d、FR-38a～FR-38c）---


def _run_weekly_scoring_and_recommendation(
    db: CloudSQLClient, llm_client, email_client, gmail_user: str, telegram_client, owner: dict, today: date
) -> None:
    """FR-37d：緊接在該週爬蟲與公司背景協作流程之後執行——不因為「這週沒有新公司背景可用」而
    跳過整次評分，只要資料庫裡已有背景資料的職缺，都會被重新納入這次評分範圍。

    比照 FR-35a「沒有新公司就整段跳過、不寄信」的既有慣例：`list_scorable_jobs()` 目前完全沒有
    可評分職缺、或評分完後兩種排名（全庫／本週新職缺）都是空清單時，直接跳過寄信與通知，避免
    寄出一封完全空白的推薦信。
    """
    scorable_jobs = list_scorable_jobs(db)
    if not scorable_jobs:
        return

    companies_by_id = _companies_by_id(db)
    profile = {
        "resume": owner.get("job_resume"),
        "years_of_experience": owner.get("years_of_experience"),
        "expected_salary_min": owner.get("expected_salary_min"),
        "expected_salary_max": owner.get("expected_salary_max"),
        "expectation": owner.get("job_expectation"),
    }
    scores = score_jobs(llm_client, profile, scorable_jobs, companies_by_id)
    if scores:
        apply_scores(db, scores)

    all_ranked = build_ranked_jobs(db, "all", companies_by_id=companies_by_id, today=today)
    new_ranked = build_ranked_jobs(db, "new_this_week", companies_by_id=companies_by_id, today=today)
    if not all_ranked and not new_ranked:
        return

    xlsx_bytes = build_job_recommendation_excel(all_ranked, new_ranked)
    send_job_recommendation_email(email_client, gmail_user, today, xlsx_bytes)
    telegram_client.send_text(
        chat_id=owner["telegram_user_id"], text=RECOMMENDATION_EMAIL_SENT_NOTIFICATION_TEXT
    )


def check_and_run_weekly_job_search(
    db: CloudSQLClient,
    job104_client,
    email_client,
    gmail_user: str,
    telegram_client,
    llm_client=None,
    now: datetime | None = None,
) -> None:
    """`/healthz` 每 10 分鐘觸發一次的其中一項排程檢查，固定台灣時間週一 08:00 這個小時內執行，
    同一天最多執行一次（`users.job_search_last_run_on` 去重，比照 `youtube_last_run_on`／
    `toeic_pipeline_last_run_on` 既有慣例）；`job_search` 功能開關關閉時跳過，不消耗任何額度。

    流程：
    1. 呼叫 `crawl_and_upsert_jobs()` 依 FR-33 各組搜尋條件爬取職缺（FR-34a～FR-34d）
    2. FR-35a：若這批職缺涉及資料庫裡還沒有背景資料的新公司（`new_company_ids` 非空），
       組 CSV 寄信給 Robin（FR-35b）、寄信成功後私訊告知（FR-35c）；若沒有新公司，
       這一步整段跳過，不寄信也不通知
    3. FR-37d：緊接著跑 Gemini 批次契合度評分（FR-37）＋雙重排名（FR-38a）＋Excel 寄送
       （FR-38b～FR-38c），見 `_run_weekly_scoring_and_recommendation()`；`llm_client` 為
       `None`（例如環境變數未設定完整）時整段跳過，不影響本函式其餘流程
    4. 更新 `job_search_last_run_on` 去重欄位

    執行過程若拋出例外，記錄警告日誌並優雅結束，不影響 `/healthz`（呼叫端 `main.py` 已經包了
    一層 try/except，這裡不重複再包，維持跟其餘 `check_and_push_*` 函式一致的分工：本函式只
    負責業務邏輯本身）。
    """
    now_local = now or _now()
    if now_local.weekday() != _WEEKLY_CRAWL_WEEKDAY or now_local.hour != _WEEKLY_CRAWL_HOUR:
        return

    owner = db.select("users", where="is_owner = %s", params=(True,), fetch_one=True)
    if owner is None:
        return
    if not toggles.is_feature_enabled(db, owner["id"], _FEATURE_KEY):
        return

    today = now_local.date()
    if owner.get("job_search_last_run_on") == today:
        return

    result = crawl_and_upsert_jobs(db, job104_client, owner["id"], now=now_local)

    if result["new_company_ids"]:
        companies = get_companies_by_ids(db, result["new_company_ids"])
        csv_text = build_new_companies_csv(companies)
        send_new_companies_email(email_client, gmail_user, today, csv_text)
        telegram_client.send_text(chat_id=owner["telegram_user_id"], text=EMAIL_SENT_NOTIFICATION_TEXT)

    if llm_client is not None:
        _run_weekly_scoring_and_recommendation(db, llm_client, email_client, gmail_user, telegram_client, owner, today)

    db.update("users", {"job_search_last_run_on": today}, where="id = %s", params=(owner["id"],))


# --- FR-40：外部管道職缺（LinkedIn／Cake 等，見 ADR-27）---

# FR-40a：外部職缺沒有 104 官方 ID，系統配發合成識別碼，格式 EXT-<內部序號>。
_EXTERNAL_ID_PREFIX = "EXT"


def _generate_external_id(internal_id: int) -> str:
    return f"{_EXTERNAL_ID_PREFIX}-{internal_id}"


def add_external_job(
    db: CloudSQLClient,
    source: str,
    title: str,
    company_name: str,
    url: str,
    content: str,
    company_background: str,
    now: datetime | None = None,
) -> str:
    """FR-40a：新增一筆非 104 來源（LinkedIn／Cake 等）的職缺，與 104 職缺共用
    `job_postings`／`job_companies`，用 `source` 欄位區分來源，取代原提案的獨立表設計（見
    ADR-27 決策 5）。統一表之後，這筆資料自動符合 FR-37／FR-38a 既有評分/排名邏輯的資料形狀
    （只要 `content`／`background` 已填），下週排程會自然把它納入評分與推薦 Excel，不需要另外
    開發一套獨立批次流程。

    沒有 104 官方 ID，合成識別碼（`EXT-<內部序號>`）的值需要 `INSERT` 後才知道的內部序號，
    採「先用暫時的唯一佔位值寫入 → 用回傳的序號組出真正的 ID → `UPDATE` 回填」兩步驟做法；
    公司與職缺的合成 ID 各自獨立配發（序號來源不同），不會互相衝突。回傳分配到的
    `job_id_104`，供 Robin 之後查詢／更新應徵狀態（FR-39）使用。
    """
    now = now or _now()

    placeholder_company_id = f"_pending_company_{uuid.uuid4().hex}"
    company_internal_id = db.insert(
        "job_companies",
        {
            "company_id_104": placeholder_company_id,
            "company_name": company_name,
            "region": None,
            "industry": None,
            "background": company_background,
            "source": source,
        },
    )
    company_id_104 = _generate_external_id(company_internal_id)
    db.update("job_companies", {"company_id_104": company_id_104}, where="id = %s", params=(company_internal_id,))

    placeholder_job_id = f"_pending_job_{uuid.uuid4().hex}"
    job_internal_id = db.insert(
        "job_postings",
        {
            "job_id_104": placeholder_job_id,
            "company_id_104": company_id_104,
            "title": title,
            "region": None,
            "url": url,
            "content": content,
            "required_years_experience": None,
            "applicant_count": None,
            "source_updated_at": None,
            "first_seen_at": now,
            "last_crawled_at": now,
            "is_closed": False,
            "source": source,
        },
    )
    job_id_104 = _generate_external_id(job_internal_id)
    db.update("job_postings", {"job_id_104": job_id_104}, where="id = %s", params=(job_internal_id,))
    return job_id_104


# --- FR-39：應徵成效追蹤（見 ADR-27）---

_APPLICATION_STATUS_APPLIED = "applied"
_APPLICATION_STATUS_INTERVIEW = "interview"
_APPLICATION_STATUS_OFFER = "offer"
_APPLICATION_STATUS_REJECTED = "rejected"

# FR-39b：任意狀態可直接設定，不強制順序；顯示用中文標籤。
_APPLICATION_STATUS_LABELS = {
    _APPLICATION_STATUS_APPLIED: "已應徵",
    _APPLICATION_STATUS_INTERVIEW: "已獲得面試",
    _APPLICATION_STATUS_OFFER: "已拿到 Offer",
    _APPLICATION_STATUS_REJECTED: "未錄取／已婉拒",
}

# FR-39b：Robin 在 Telegram 打的中文語句關鍵字（已去除空白）→ 內部狀態值對照表；「未錄取」
# 「已婉拒」語意相近，同時收兩種講法降低使用摩擦，不強迫 Robin 只能用固定字眼。呼叫端（見
# router.py `_APPLICATION_STATUS_PATTERN`）比對前需先把擷取到的文字空白去除再查表（「已拿到
# Offer」中間可能有 0～多個空白）。
APPLICATION_STATUS_TEXT_TO_STATUS = {
    "已應徵": _APPLICATION_STATUS_APPLIED,
    "已獲得面試": _APPLICATION_STATUS_INTERVIEW,
    "已拿到Offer": _APPLICATION_STATUS_OFFER,
    "已婉拒": _APPLICATION_STATUS_REJECTED,
    "未錄取": _APPLICATION_STATUS_REJECTED,
}


def application_status_label(status: str) -> str:
    """把內部狀態值轉成中文標籤，供 `router.py` 組成功回覆訊息使用（不直接存取模組私有的
    `_APPLICATION_STATUS_LABELS`）。"""
    return _APPLICATION_STATUS_LABELS.get(status, status)


def record_application_status(
    db: CloudSQLClient, job_id_104: str, status: str, now: datetime | None = None
) -> bool:
    """FR-39b／FR-39c：把應徵狀態變化寫入 `job_applications` 歷程表（append-only，不覆蓋既有
    紀錄，每次變更各自一筆＋時間戳）。寫入前先確認 `job_id_104` 存在於 `job_postings`——ID
    打錯字不會憑空建立孤兒紀錄，回傳 `False` 讓呼叫端可以告知 Robin「找不到這個職缺 ID」。
    """
    job = db.select("job_postings", where="job_id_104 = %s", params=(job_id_104,), fetch_one=True)
    if job is None:
        return False
    db.insert("job_applications", {"job_id_104": job_id_104, "status": status, "created_at": now or _now()})
    return True


def list_latest_application_statuses(db: CloudSQLClient) -> list[dict]:
    """FR-39（追加，供「我的應徵紀錄」查詢指令使用）：撈出每個職缺目前最新的應徵狀態，依最新
    更新時間由新到舊排序。`job_applications` 為 append-only 歷程表，同一 `job_id_104` 可能有
    多筆歷史紀錄，這裡只取 `created_at` 最新的一筆視為「目前狀態」。
    """
    all_records = db.select("job_applications")
    latest_by_job: dict[str, dict] = {}
    for record in all_records:
        job_id = record["job_id_104"]
        if job_id not in latest_by_job or record["created_at"] > latest_by_job[job_id]["created_at"]:
            latest_by_job[job_id] = record

    jobs_by_id = {j["job_id_104"]: j for j in db.select("job_postings")}
    results = []
    for job_id, record in latest_by_job.items():
        job = jobs_by_id.get(job_id, {})
        results.append(
            {
                "job_id_104": job_id,
                "title": job.get("title", "（職缺資料不存在）"),
                "status": record["status"],
                "updated_at": record["created_at"],
            }
        )
    results.sort(key=lambda r: r["updated_at"], reverse=True)
    return results


def format_application_statuses(statuses: list[dict]) -> str:
    """把 `list_latest_application_statuses()` 的結果組成文字清單，供「我的應徵紀錄」查詢指令
    使用。"""
    if not statuses:
        return "目前還沒有任何應徵紀錄喔！用「ID=XXX 職缺已應徵」這類語句就可以開始記錄了。"

    lines = ["📋 目前的應徵紀錄："]
    for item in statuses:
        label = _APPLICATION_STATUS_LABELS.get(item["status"], item["status"])
        lines.append(f"・{item['title']}（ID={item['job_id_104']}）：{label}")
    return "\n".join(lines)


def get_profile(db: CloudSQLClient, user_id: int) -> dict:
    """取得求職設定所需的使用者欄位。"""
    return db.select("users", where="id = %s", params=(user_id,), fetch_one=True) or {}


def update_profile_field(db: CloudSQLClient, user_id: int, field: str, value) -> None:
    """只更新 FR-41 選單目前操作的單一履歷或必要條件欄位。"""
    allowed_fields = {
        "job_resume",
        "job_expectation",
        "years_of_experience",
        "expected_salary_min",
        "expected_salary_max",
    }
    if field not in allowed_fields:
        raise ValueError(f"不支援的求職設定欄位：{field}")
    db.update("users", {field: value}, where="id = %s", params=(user_id,))


def delete_search_criteria(db: CloudSQLClient, user_id: int, criteria_id: int) -> bool:
    """刪除指定使用者的一筆搜尋條件，避免跨帳號刪除。"""
    return bool(db.delete("job_search_criteria", "id = %s AND user_id = %s", (criteria_id, user_id)))


def list_jobs_by_score(db: CloudSQLClient) -> list[dict]:
    """職缺清單依契合度分數由高至低，未評分職缺排在最後。"""
    return sorted(
        db.select("job_postings"),
        key=lambda job: (job.get("score") is not None, job.get("score") or 0),
        reverse=True,
    )


def list_jobs_by_latest_application_status(db: CloudSQLClient, status: str) -> list[dict]:
    """依 append-only 應徵歷程找出目前處於指定狀態的職缺。"""
    latest = {item["job_id_104"]: item for item in list_latest_application_statuses(db)}
    return [item for item in latest.values() if item["status"] == status]


def set_job_closed_manually(db: CloudSQLClient, job_id_104: str, is_closed: bool) -> bool:
    """寫入人工職缺開關並保護它不被下一輪 104 爬蟲覆寫。"""
    return bool(
        db.update(
            "job_postings",
            {"is_closed": is_closed, "is_closed_manual_override": True},
            where="job_id_104 = %s",
            params=(job_id_104,),
        )
    )
